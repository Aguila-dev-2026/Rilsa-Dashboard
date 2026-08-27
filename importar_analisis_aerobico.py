"""Importa análisis de Planta Alta, Planta Aeróbica y Efluente.

La salida conserva el formato largo consumido por el dashboard y añade
metadatos para distinguir punto, turno, origen, cálculos y calificadores.
Los pesos de filtros, volúmenes de muestra, conteos y auxiliares de laboratorio
se excluyen deliberadamente.
"""

from __future__ import annotations

import argparse
import math
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


RAIZ_PROYECTO = Path(__file__).resolve().parent
ENTRADA_PREDETERMINADA = (
    RAIZ_PROYECTO / "datos" / "Análisis Planta Aeróbica.xlsx"
)
SALIDA_PREDETERMINADA = (
    RAIZ_PROYECTO / "datos_generados" / "analisis_planta_aerobica.xlsx"
)
FUENTE = "Análisis Planta Aeróbica.xlsx"

COLUMNAS_SALIDA = [
    "Fecha",
    "Area",
    "Punto",
    "Turno",
    "Parametro",
    "Valor",
    "Unidad",
    "TipoDato",
    "Calificador",
    "Fuente",
    "Hoja",
]


def convertir_medicion(valor: object) -> tuple[float | None, str]:
    """Convierte números sin confundir límites analíticos con ceros."""
    if valor is None or isinstance(valor, bool):
        return None, ""
    if isinstance(valor, (int, float)):
        numero = float(valor)
        return (numero, "") if math.isfinite(numero) else (None, "")

    texto = str(valor).strip().replace("\u00a0", "")
    if not texto or texto in {"-", "—", "–", "−"}:
        return None, ""

    coincidencia = re.match(
        r"^\s*([<>≤≥~*]*)\s*([-+]?\d+(?:[.,]\d+)?)\s*$",
        texto,
    )
    if not coincidencia:
        return None, ""

    calificador, numero_texto = coincidencia.groups()
    numero_texto = numero_texto.replace(",", ".")
    try:
        return float(numero_texto), calificador
    except ValueError:
        return None, ""


def normalizar_turno(valor: object) -> str:
    """Conserva turno, frecuencia o tipo de muestra sin fragmentar mayúsculas."""
    if valor is None:
        return ""

    original = str(valor).strip()
    if not original or original in {"-", "—", "–", "−"}:
        return ""

    texto = original.casefold()
    normalizados = {
        "mañana": "Mañana",
        "tarde": "Tarde",
        "noche": "Noche",
        "medio dia": "Medio día",
        "mediodía": "Medio día",
        "compuesta": "Compuesta",
        "compuesto": "Compuesta",
        "puntual": "Puntual",
        "contramuestra": "Contramuestra",
        "purga r1": "Purga R1",
        "lodo fondo cono (purga)": "Lodo fondo cono (purga)",
    }
    return normalizados.get(texto, original[:60])


def agregar_registro(
    registros: list[dict],
    *,
    fecha: object,
    area: str,
    punto: str,
    turno: str,
    parametro: str,
    unidad: str,
    tipo_dato: str,
    valor: object,
    hoja: str,
    prioridad: int,
) -> None:
    fecha_normalizada = pd.to_datetime(fecha, errors="coerce")
    numero, calificador = convertir_medicion(valor)
    if pd.isna(fecha_normalizada) or numero is None:
        return

    fecha_normalizada = pd.Timestamp(fecha_normalizada).normalize()
    if (
        fecha_normalizada < pd.Timestamp("2020-01-01")
        or fecha_normalizada > pd.Timestamp.today().normalize()
    ):
        return

    registros.append(
        {
            "Fecha": fecha_normalizada,
            "Area": area,
            "Punto": punto,
            "Turno": turno,
            "Parametro": parametro,
            "Valor": numero,
            "Unidad": unidad,
            "TipoDato": tipo_dato,
            "Calificador": calificador,
            "Fuente": FUENTE,
            "Hoja": hoja,
            "_Prioridad": prioridad,
        }
    )


def importar_columnas(
    hoja,
    registros: list[dict],
    *,
    fila_inicial: int,
    columna_fecha: int,
    columna_turno: int | None,
    area: str,
    punto: str,
    mapeo: dict[int, tuple[str, str, str]],
    prioridad: int,
    celdas_rotas: set[tuple[str, int, int]] | None = None,
) -> None:
    celdas_rotas = celdas_rotas or set()
    for numero_fila, fila in enumerate(
        hoja.iter_rows(min_row=fila_inicial, values_only=True),
        start=fila_inicial,
    ):
        fecha = fila[columna_fecha - 1] if len(fila) >= columna_fecha else None
        turno = (
            normalizar_turno(fila[columna_turno - 1])
            if columna_turno and len(fila) >= columna_turno
            else ""
        )
        if turno.casefold() == "tarde":
            continue

        for columna, (parametro, unidad, tipo_dato) in mapeo.items():
            if (hoja.title, numero_fila, columna) in celdas_rotas:
                continue
            valor = fila[columna - 1] if len(fila) >= columna else None
            agregar_registro(
                registros,
                fecha=fecha,
                area=area,
                punto=punto,
                turno=turno,
                parametro=parametro,
                unidad=unidad,
                tipo_dato=tipo_dato,
                valor=valor,
                hoja=hoja.title,
                prioridad=prioridad,
            )


MAPEO_AFLUENTE = {
    4: ("DQO", "mg/L", "Medición"),
    5: ("DBO estimada", "mg/L", "Estimación"),
    6: ("DQO TK2", "mg/L", "Medición"),
    7: ("Remoción DQO", "%", "Cálculo"),
    8: ("Descarga a Reactor N°1", "m3", "Medición"),
    9: ("Carga DBO5", "kg/día", "Cálculo"),
    10: ("Máxima descarga diaria", "m3/día", "Cálculo"),
    11: ("pH", "", "Medición"),
    12: ("Temperatura", "°C", "Medición"),
    13: ("Conductividad", "mS/cm", "Medición"),
    14: ("Relación C:P", "", "Cálculo"),
    15: ("Relación N:P", "", "Cálculo"),
    16: ("Volumen de alimentación", "m3", "Medición"),
    17: ("Nitrógeno total", "mg/L", "Medición"),
    18: ("Carga de nitrógeno total", "kg/día", "Cálculo"),
    19: ("NH3", "mg/L", "Medición"),
    20: ("NO3-", "mg/L", "Medición"),
    21: ("Fósforo total", "mg/L", "Medición"),
    22: ("Carga de fósforo total", "kg/día", "Cálculo"),
    23: ("PO4-3", "mg/L", "Medición"),
    24: ("P2O5", "mg/L", "Medición"),
    25: ("Color aparente", "PCU", "Medición"),
    26: ("Turbidez", "FAU", "Medición"),
    27: ("Sulfato", "mg/L", "Medición"),
    32: ("SST", "mg/L", "Cálculo"),
    33: ("SSV", "mg/L", "Cálculo"),
    34: ("SSF", "mg/L", "Cálculo"),
}

MAPEO_REACTOR_1 = {
    3: ("pH", "", "Medición"),
    4: ("Temperatura", "°C", "Medición"),
    5: ("Conductividad", "mS/cm", "Medición"),
    10: ("SST", "mg/L", "Cálculo"),
    11: ("SSV", "mg/L", "Cálculo"),
    12: ("SSF", "mg/L", "Cálculo"),
    13: ("Sólidos sedimentables 30 min", "mL/L", "Medición"),
    14: ("Sólidos sedimentables 60 min", "mL/L", "Medición"),
    15: ("IVL", "mL/g", "Cálculo"),
    16: ("Edad del lodo", "días", "Medición"),
    17: ("DQO clarificado", "mg/L", "Medición"),
    29: ("Nitrógeno total", "mg/L", "Medición"),
    30: ("NH3", "mg/L", "Medición"),
    31: ("NO3-", "mg/L", "Medición"),
    32: ("Fósforo total", "mg/L", "Medición"),
    33: ("PO4-3", "mg/L", "Medición"),
    34: ("P2O5", "mg/L", "Medición"),
    27: ("Relación C:P", "", "Cálculo"),
    28: ("Relación N:P", "", "Cálculo"),
    35: ("Relación F/M", "", "Cálculo"),
}

MAPEO_CLARIFICADO_R1 = {
    18: ("Turbidez", "FAU", "Medición"),
    23: ("SST", "mg/L", "Cálculo"),
    24: ("SSV", "mg/L", "Cálculo"),
    25: ("SSF", "mg/L", "Cálculo"),
}

MAPEO_EFLUENTE = {
    3: ("DQO", "mg/L", "Medición"),
    4: ("DBO estimada", "mg/L", "Estimación"),
    5: ("pH", "", "Medición"),
    6: ("Temperatura", "°C", "Medición"),
    7: ("Conductividad", "mS/cm", "Medición"),
    8: ("Color aparente", "PCU", "Medición"),
    9: ("Turbidez", "FAU", "Medición"),
    10: ("Sulfato", "mg/L", "Medición"),
    11: ("Boro total", "mg/L", "Medición"),
    12: ("Nitrógeno total", "mg/L", "Medición"),
    13: ("NH3", "mg/L", "Medición"),
    14: ("Nitrógeno", "mg/L", "Medición"),
    19: ("SST", "mg/L", "Cálculo"),
    20: ("SSV", "mg/L", "Cálculo"),
    21: ("SSF", "mg/L", "Cálculo"),
    25: ("Sólidos disueltos", "mg/L", "Cálculo"),
    29: ("Cloruro", "mg/L", "Cálculo"),
    30: ("Cloro libre", "ppm", "Medición"),
    31: ("Descarga a riego", "Sí/No", "Registro"),
    32: ("Descarga a infiltración", "Sí/No", "Registro"),
}


def mapeo_bloque_solidos(base: int) -> dict[int, tuple[str, str, str]]:
    return {
        base: ("pH", "", "Medición"),
        base + 1: ("Temperatura", "°C", "Medición"),
        base + 2: ("Conductividad", "mS/cm", "Medición"),
        base + 6: ("SST", "mg/L", "Cálculo"),
        base + 7: ("SSV", "mg/L", "Cálculo"),
        base + 8: ("SSF", "mg/L", "Cálculo"),
        base + 9: ("Relación SSF/SSV", "", "Cálculo"),
        base + 10: ("Sólidos sedimentables 30 min", "mL/L", "Medición"),
        base + 11: ("IVL", "mL/g", "Cálculo"),
        base + 12: ("Sólidos sedimentables 60 min", "mL/L", "Medición"),
    }


def importar_analisis_aerobico(
    ruta_entrada: Path = ENTRADA_PREDETERMINADA,
    ruta_salida: Path = SALIDA_PREDETERMINADA,
) -> pd.DataFrame:
    if not ruta_entrada.exists():
        raise FileNotFoundError(f"No existe la planilla de origen: {ruta_entrada}")

    libro = load_workbook(ruta_entrada, data_only=True, read_only=True)
    libro_formulas = load_workbook(ruta_entrada, data_only=False, read_only=True)
    requeridas = {
        "AFLUENTE",
        "REACTOR N°1 Y DESCARGA",
        "PLANTA AEROBICA BAJA",
        "CLARIFICADOR BIOLÓGICO",
        "EFLUENTE",
    }
    faltantes = requeridas.difference(libro.sheetnames)
    if faltantes:
        raise ValueError(f"Faltan hojas requeridas: {', '.join(sorted(faltantes))}")

    celdas_rotas: set[tuple[str, int, int]] = set()
    for nombre_hoja in requeridas:
        for fila in libro_formulas[nombre_hoja].iter_rows():
            for celda in fila:
                if isinstance(celda.value, str) and "#REF!" in celda.value:
                    celdas_rotas.add((nombre_hoja, celda.row, celda.column))

    registros: list[dict] = []

    importar_columnas(
        libro["AFLUENTE"],
        registros,
        fila_inicial=7,
        columna_fecha=2,
        columna_turno=3,
        area="Planta Alta",
        punto="Afluente",
        mapeo=MAPEO_AFLUENTE,
        prioridad=400,
    )
    importar_columnas(
        libro["REACTOR N°1 Y DESCARGA"],
        registros,
        fila_inicial=7,
        columna_fecha=1,
        columna_turno=2,
        area="Planta Aeróbica",
        punto="Reactor N°1",
        mapeo=MAPEO_REACTOR_1,
        prioridad=400,
    )
    importar_columnas(
        libro["REACTOR N°1 Y DESCARGA"],
        registros,
        fila_inicial=7,
        columna_fecha=1,
        columna_turno=2,
        area="Planta Aeróbica",
        punto="Clarificado Reactor N°1",
        mapeo=MAPEO_CLARIFICADO_R1,
        prioridad=400,
    )

    hoja_pab = libro["PLANTA AEROBICA BAJA"]
    for punto, base in {
        "Reactor N°2": 4,
        "Reactor N°3": 18,
        "Reactor N°4": 32,
        "Reactor N°5": 46,
        "Digestor": 60,
        "Recirculación": 74,
    }.items():
        importar_columnas(
            hoja_pab,
            registros,
            fila_inicial=6,
            columna_fecha=1,
            columna_turno=None,
            area="Planta Aeróbica",
            punto=punto,
            mapeo=mapeo_bloque_solidos(base),
            prioridad=400,
        )

    importar_columnas(
        hoja_pab,
        registros,
        fila_inicial=6,
        columna_fecha=1,
        columna_turno=None,
        area="Planta Aeróbica",
        punto="Clarificador Biológico",
        mapeo={
            89: ("pH", "", "Medición"),
            90: ("Temperatura", "°C", "Medición"),
            91: ("Conductividad", "mS/cm", "Medición"),
            92: ("Turbidez", "FAU", "Medición"),
            96: ("SST", "mg/L", "Cálculo"),
            97: ("SSV", "mg/L", "Cálculo"),
            98: ("SSF", "mg/L", "Cálculo"),
            99: ("Relación SSF/SSV", "", "Cálculo"),
        },
        prioridad=400,
    )

    importar_columnas(
        libro["CLARIFICADOR BIOLÓGICO"],
        registros,
        fila_inicial=6,
        columna_fecha=1,
        columna_turno=None,
        area="Planta Aeróbica",
        punto="Clarificador Biológico",
        mapeo={
            2: ("pH", "", "Medición"),
            3: ("Temperatura", "°C", "Medición"),
            4: ("Conductividad", "mS/cm", "Medición"),
            5: ("Turbidez", "FAU", "Medición"),
            6: ("DQO", "mg/L", "Medición"),
            7: ("Sólidos sedimentables 30 min", "mL/L", "Medición"),
        },
        prioridad=500,
    )

    importar_columnas(
        libro["EFLUENTE"],
        registros,
        fila_inicial=6,
        columna_fecha=1,
        columna_turno=2,
        area="Efluente",
        punto="Efluente",
        mapeo=MAPEO_EFLUENTE,
        prioridad=400,
        celdas_rotas=celdas_rotas,
    )

    if not registros:
        raise ValueError("No se encontraron mediciones válidas para importar.")

    resultado = pd.DataFrame(registros)
    claves = ["Fecha", "Area", "Punto", "Turno", "Parametro"]
    resultado = (
        resultado.sort_values([*claves, "_Prioridad"])
        .drop_duplicates(subset=claves, keep="last")
        .drop(columns="_Prioridad")
        .sort_values(["Fecha", "Area", "Punto", "Turno", "Parametro"])
        .reset_index(drop=True)
    )
    resultado = resultado[COLUMNAS_SALIDA]

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    resultado.to_excel(ruta_salida, index=False)

    print(f"Archivo de origen: {ruta_entrada.name}")
    print(f"Registros importados: {len(resultado)}")
    print(f"Áreas: {', '.join(sorted(resultado['Area'].unique()))}")
    print(f"Archivo generado: {ruta_salida}")
    return resultado


def argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Importa análisis de Planta Alta, Planta Aeróbica y Efluente."
    )
    parser.add_argument("--entrada", type=Path, default=ENTRADA_PREDETERMINADA)
    parser.add_argument("--salida", type=Path, default=SALIDA_PREDETERMINADA)
    return parser.parse_args()


if __name__ == "__main__":
    opciones = argumentos()
    importar_analisis_aerobico(opciones.entrada, opciones.salida)
